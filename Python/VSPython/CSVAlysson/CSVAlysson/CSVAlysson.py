import tkinter as tk
from tkinter import filedialog
import codecs
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def remove_blank_lines(file_path):
    # Lê o conteúdo do arquivo
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Remove as linhas em branco
    non_blank_lines = [line for line in lines if line.strip()]

    # Escreve o conteúdo processado em um novo arquivo
    current_date = datetime.now().strftime("%m-%Y")
    output_file_name = f'arquivoLimpo_{current_date}.txt'
    with open(output_file_name, 'w') as file:
        file.writelines(non_blank_lines)


def process_file(file_path):
    global total_line
    if file_path:
        # Cria um novo arquivo Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Largura de cada coluna
        data_column = worksheet.column_dimensions['A']
        n_ordem_column = worksheet.column_dimensions['B']
        material_column = worksheet.column_dimensions['C']
        quantidade_column = worksheet.column_dimensions['D']
        unidade_column = worksheet.column_dimensions['E']
        valor_column = worksheet.column_dimensions['F']

        # Define a largura mínima das colunas
        data_column.width = 17  # Coluna A
        n_ordem_column.width = 12  # Coluna B
        material_column.width = 80  # Coluna C
        quantidade_column.width = 16  # Coluna D
        unidade_column.width = 11  # Coluna E
        valor_column.width = 15  # Coluna F

        # Cria as colunas do cabeçalho
        header_columns = ['Data', 'N.Ordem', 'Material', 'Quantidade', 'Unidade', 'Valor']
        for col_num, column_text in enumerate(header_columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_text)
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # ...

        # Especifique a codificação correta do arquivo
        encoding = 'UTF-8'  # Ou outra codificação adequada

        # Abra o arquivo para leitura usando a codificação correta
        with codecs.open(file_path, 'r', encoding) as file:
            lines = file.readlines()

        processed_lines = []
        foi = 0
        templine = ''
        total = 0
        previous_vehicle_total = 0
        previous_vehicle_line = ''
        for line in lines:
            if "VEICULO: " in line and any(c.isdigit() for c in line.split("VEICULO: ")[1].split(" ")[0].strip()):
                if foi != 0:
                    # Adiciona a linha com o total do veículo anterior
                    if previous_vehicle_total > 0:
                        previous_vehicle_line = previous_vehicle_line.rstrip()
                        total_line = ' TOTAL DO ' + previous_vehicle_line + f': R$ {previous_vehicle_total:.2f}\n'
                        processed_lines.append(total_line)
                        branco = '\n'
                        processed_lines.append(branco)

                        # Adiciona no Excel
                        columns = [total_line.strip()]
                        worksheet.append(columns)



                # Reinicia o total para o novo veículo
                total = 0
                foi = 1
                templine = line
                veiculo_anterior = line.split("VEICULO: ")[1].split(" ")[0].strip()
                templine = templine.rstrip()
                branco = '\n'
                processed_lines.append(templine)
                processed_lines.append(branco)
                previous_vehicle_line = templine
                previous_vehicle_total = 0
                continue

            if not line.strip() or not line.strip()[
                0].isdigit() or "   OLEO" in line or "   ARLA" in line or "000" in line:
                continue

            # Dividir a linha nos campos correspondentes
            data = line[:10].strip()
            n_ordem = line[11:22].strip()
            material = line[23:100].strip()
            quantidade = line[100:113].strip()
            unidade = line[113:121].strip()  # Aumentei o tamanho para acomodar a unidade corretamente
            valor = line[122:].strip()  # Corrigi o índice do valor

            # Adicionar os campos processados à lista
            processed_lines.append(f"{data}\t{n_ordem}\t{material}\t{quantidade}\t{unidade}\t{valor}\n")

            # Extrair o valor da linha
            value_str = line.strip().split()[-1].replace('.', '').replace(',', '.')

            # Check if value_str is a valid float
            if value_str.replace('.', '', 1).isdigit():
                value = float(value_str)
                total += value
                previous_vehicle_total += value

        # Adiciona o total do último veículo
        if total > 0:
            templine = templine.rstrip()
            total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}'
            processed_lines.append(total_line)
            processed_lines.append('\n')
            columns = total_line.strip().split('\t')
            worksheet.append(columns)

        processed_text = ''.join(processed_lines)

        with open('arquivo_processado.txt', 'w') as file:
            file.write(processed_text)

        # Remove as linhas em branco do arquivo processado
        remove_blank_lines('arquivo_processado.txt')

        with open('arquivo_processado.txt', 'r') as file:
            processed_lines = file.readlines()

        row_num = 2
        total = 0  # Variável para armazenar a soma dos valores
        for line in processed_lines:
            columns = line.strip().split('\t')
            for col_num, column_text in enumerate(columns, start=1):
                worksheet.cell(row=row_num, column=col_num, value=column_text)
            value_str = columns[-1].replace('.', '').replace(',', '.')  # Valor como string
            if value_str.replace('.', '', 1).isdigit():
                value = float(value_str)  # Converte o valor para float
                total += value  # Adiciona o valor à soma total
            row_num += 1

        # Adiciona a soma total à célula "Valor" na última linha
        total_row = worksheet.max_row + 1
        worksheet.cell(row=total_row, column=header_columns.index("Unidade") + 1, value="TOTAL")
        worksheet.cell(row=total_row, column=header_columns.index("Valor") + 1, value=total)
        total_value_cell = worksheet.cell(row=total_row, column=header_columns.index("Valor") + 2)
        total_value_cell.number_format = '#,##0.00'  # Formatação do número com duas casas decimais

        # Estiliza a célula "TOTAL"
        total_cell = worksheet.cell(row=total_row, column=header_columns.index("Unidade") + 1)
        total_cell.font = Font(name='Arial', size=12, bold=True)
        total_cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Salva o arquivo Excel
        current_date = datetime.now().strftime("%m-%Y")
        output_file_name = f'Arquivo_Final_{current_date}.xlsx'
        workbook.save(output_file_name)


def select_file():
    # Abre a janela de seleção de arquivo
    file_path = filedialog.askopenfilename()
    if file_path:
        # Chama a função para processar o arquivo selecionado
        process_file(file_path)

        # Mostra uma mensagem de conclusão
        tk.messagebox.showinfo("Concluído", "Arquivo processado e salvo com sucesso como Excel!")
    else:
        tk.messagebox.showinfo("Sem Arquivo", "NENHUM ARQUIVO SELECIONADO!")


# Cria uma janela principal
root = tk.Tk()
root.title("Limpador de Arquivos")

# Define o ícone da janela
icon = tk.PhotoImage(file="dependencias/icone.png")
root.iconphoto(True, icon)

# Define o tamanho mínimo da janela
root.minsize(400, 300)

# Cria o frame para centralizar o botão
frame = tk.Frame(root)
frame.pack(expand=True)

# Cria as instruções
instructions = tk.Label(frame, text='Selecione um arquivo para processar, o arquivo terá saída em excel com nome de "Arquivo_Final_DATA".')
instructions.pack(pady=10)

# Cria o botão
button = tk.Button(frame, text="Selecionar Arquivo", command=select_file)
button.pack(pady=10)

# Cria a assinatura
signature = tk.Label(root, text="Pensado e desenvolvido por Fábio H. Araújo - contato@fabioharaujo.com.br",
                     font=("Arial", 8), foreground="gray")
signature.pack(side="bottom", pady=5, padx=10, anchor="se")


root.mainloop()
