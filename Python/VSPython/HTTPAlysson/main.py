import os
from functools import wraps
from flask import Flask, render_template, request, send_file, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
import codecs
import tempfile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = '@Server1808'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://alyshow:alyshow1463@localhost/alyshow'  # Configura o URI do banco de dados
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Desativa o rastreamento de modificações
db = SQLAlchemy(app)



@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']

    if file:
        # Salva o arquivo temporariamente
        temp_file = tempfile.NamedTemporaryFile(delete=False)
        file.save(temp_file.name)
        temp_file.close()

        # Processa o arquivo
        process_file(temp_file.name)

        # Remove o arquivo temporário
        # Comente a linha abaixo se quiser manter o arquivo temporário para fins de depuração
        # os.remove(temp_file.name)

        return 'Arquivo processado e salvo com sucesso como Excel!'

    return 'Nenhum arquivo recebido.'


@app.route('/download')
def download():
    current_date = datetime.now().strftime("%m-%Y")
    file_path = f'Arquivo_Final_{current_date}.xlsx'
    return send_file(file_path, as_attachment=True)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True)
    password = db.Column(db.String(50))

    def __init__(self, username, password):
        self.username = username
        self.password = password


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = User.query.filter_by(username=username, password=password).first()
        if users:
            session['logged_in'] = True  # Define a sessão como logada
            return redirect(url_for('index'))
        else:
            return render_template('login.html', message='Usuário ou senha inválidos')
    return render_template('login.html', message='')


def remove_blank_lines(file_path):
    # Lê o conteúdo do arquivo
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Remove as linhas em branco
    non_blank_lines = [line for line in lines if line.strip()]

    # Escreve o conteúdo processado em um novo arquivo
    current_date = datetime.now().strftime("%m-%Y")
    output_file_name = f'arquivoLimpo_{current_date}.txt'
    with open(output_file_name, 'w') as file:
        file.writelines(non_blank_lines)


def remove_blank_lines(file_path):
    # Lê o conteúdo do arquivo
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Remove as linhas em branco
    non_blank_lines = [line for line in lines if line.strip()]

    # Escreve o conteúdo processado em um novo arquivo
    current_date = datetime.now().strftime("%m-%Y")
    output_file_name = f'arquivoLimpo_{current_date}.txt'
    with open(output_file_name, 'w') as file:
        file.writelines(non_blank_lines)

def process_file(file_path):
    global total_line
    if file_path:
        # Cria um novo arquivo Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Largura de cada coluna
        data_column = worksheet.column_dimensions['A']
        n_ordem_column = worksheet.column_dimensions['B']
        material_column = worksheet.column_dimensions['C']
        quantidade_column = worksheet.column_dimensions['D']
        unidade_column = worksheet.column_dimensions['E']
        valor_column = worksheet.column_dimensions['F']

        # Define a largura mínima das colunas
        data_column.width = 17  # Coluna A
        n_ordem_column.width = 12  # Coluna B
        material_column.width = 80  # Coluna C
        quantidade_column.width = 16  # Coluna D
        unidade_column.width = 11  # Coluna E
        valor_column.width = 15  # Coluna F

        # Cria as colunas do cabeçalho
        header_columns = ['Data', 'N.Ordem', 'Material', 'Quantidade', 'Unidade', 'Valor']
        for col_num, column_text in enumerate(header_columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_text)
            cell.font = Font(name='Arial', size=12, bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # ...

        # Especifique a codificação correta do arquivo
        encoding = 'UTF-8'  # Ou outra codificação adequada

        # Abra o arquivo para leitura usando a codificação correta
        with codecs.open(file_path, 'r', encoding) as file:
            lines = file.readlines()

        processed_lines = []
        foi = 0
        templine = ''
        total = 0
        previous_vehicle_total = 0
        previous_vehicle_line = ''
        for line in lines:
            if "VEICULO: " in line and any(c.isdigit() for c in line.split("VEICULO: ")[1].split(" ")[0].strip()):
                if foi != 0:
                    # Adiciona a linha com o total do veículo anterior
                    if previous_vehicle_total > 0:
                        previous_vehicle_line = previous_vehicle_line.rstrip()
                        total_line = ' TOTAL DO ' + previous_vehicle_line + f': R$ {previous_vehicle_total:.2f}\n'
                        processed_lines.append(total_line)
                        branco = '\n'
                        processed_lines.append(branco)

                        # Adiciona no Excel
                        columns = [total_line.strip()]
                        worksheet.append(columns)



                # Reinicia o total para o novo veículo
                total = 0
                foi = 1
                templine = line
                veiculo_anterior = line.split("VEICULO: ")[1].split(" ")[0].strip()
                templine = templine.rstrip()
                branco = '\n'
                processed_lines.append(templine)
                processed_lines.append(branco)
                previous_vehicle_line = templine
                previous_vehicle_total = 0
                continue

            if not line.strip() or not line.strip()[
                0].isdigit() or "   OLEO" in line or "   ARLA" in line or "000" in line:
                continue

            # Dividir a linha nos campos correspondentes
            data = line[:10].strip()
            n_ordem = line[11:22].strip()
            material = line[23:100].strip()
            quantidade = line[100:113].strip()
            unidade = line[113:121].strip()  # Aumentei o tamanho para acomodar a unidade corretamente
            valor = line[122:].strip()  # Corrigi o índice do valor

            # Adicionar os campos processados à lista
            processed_lines.append(f"{data}\t{n_ordem}\t{material}\t{quantidade}\t{unidade}\t{valor}\n")

            # Extrair o valor da linha
            value_str = line.strip().split()[-1].replace('.', '').replace(',', '.')

            # Check if value_str is a valid float
            if value_str.replace('.', '', 1).isdigit():
                value = float(value_str)
                total += value
                previous_vehicle_total += value

        # Adiciona o total do último veículo
        if total > 0:
            templine = templine.rstrip()
            total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}'
            processed_lines.append(total_line)
            processed_lines.append('\n')
            columns = total_line.strip().split('\t')
            worksheet.append(columns)

        processed_text = ''.join(processed_lines)

        with open('arquivo_processado.txt', 'w') as file:
            file.write(processed_text)

        # Remove as linhas em branco do arquivo processado
        remove_blank_lines('arquivo_processado.txt')

        with open('arquivo_processado.txt', 'r') as file:
            processed_lines = file.readlines()

        row_num = 2
        total = 0  # Variável para armazenar a soma dos valores
        for line in processed_lines:
            columns = line.strip().split('\t')
            for col_num, column_text in enumerate(columns, start=1):
                worksheet.cell(row=row_num, column=col_num, value=column_text)
            value_str = columns[-1].replace('.', '').replace(',', '.')  # Valor como string
            if value_str.replace('.', '', 1).isdigit():
                value = float(value_str)  # Converte o valor para float
                total += value  # Adiciona o valor à soma total
            row_num += 1

        # Adiciona a soma total à célula "Valor" na última linha
        total_row = worksheet.max_row + 1
        worksheet.cell(row=total_row, column=header_columns.index("Unidade") + 1, value="TOTAL")
        worksheet.cell(row=total_row, column=header_columns.index("Valor") + 1, value=total)
        total_value_cell = worksheet.cell(row=total_row, column=header_columns.index("Valor") + 2)
        total_value_cell.number_format = '#,##0.00'  # Formatação do número com duas casas decimais

        # Estiliza a célula "TOTAL"
        total_cell = worksheet.cell(row=total_row, column=header_columns.index("Unidade") + 1)
        total_cell.font = Font(name='Arial', size=12, bold=True)
        total_cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Salva o arquivo Excel
        current_date = datetime.now().strftime("%m-%Y")
        output_file_name = f'Arquivo_Final_{current_date}.xlsx'
        workbook.save(output_file_name)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Verifica se o usuário está logado
        if not session.get('logged_in'):
            return redirect('/login')  # Redireciona para a tela de login
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process():
    file = request.files['file']
    if file:
        # Salva o arquivo em disco temporariamente
        file_path = 'temp_file.txt'
        file.save(file_path)

        # Chama a função para processar o arquivo selecionado
        process_file(file_path)

        # Remove o arquivo temporário
        os.remove(file_path)

        # Remove as linhas em branco do arquivo processado
        remove_blank_lines('arquivo_processado.txt')

        # Restante do código...

        # Retorna o arquivo processado para download
        return send_file('Arquivo_Final_DATA.xlsx', as_attachment=True, attachment_filename='Arquivo_Final_DATA.xlsx')

    return 'Nenhum arquivo selecionado.'

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8343)
    app.run(debug=True)
