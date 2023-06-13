import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
import codecs
from datetime import datetime

def remove_blank_lines(file_path):
    # Lê o conteúdo do arquivo
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Remove as linhas em branco
    non_blank_lines = [line for line in lines if line.strip()]

    # Escreve o conteúdo processado em um novo arquivo
    current_date = datetime.now().strftime("%m-%Y")
    output_file_name = f'arquivoLimpo_{current_date}.txt'
    with open(output_file_name, 'w') as file:
        file.writelines(non_blank_lines)

def process_file(file_path):
    if file_path:
        # Especifique a codificação correta do arquivo
        encoding = 'UTF-8'  # Ou outra codificação adequada

        # Abra o arquivo para leitura usando a codificação correta
        with codecs.open(file_path, 'r', encoding) as file:
            lines = file.readlines()

        processed_lines = []

        # Adiciona o cabeçalho
        header =  " MÓDULO CONTROLE DE ESTOQUE                                  ORDENS DE SERVICO DE VEICULOS                           PÁGINA :          1\n"
        header += " 000-TODAS AS EMPRESAS                                       Emissão: 01/03/2023 A 31/05/2023                        DATA   : 07/06/2023\n"
        header += " ---------------------------------------------------------------------------------------------------------------------------------------\n"
        header += " MOTORISTA:                                                                                                                             \n"
        header += " VEICULO: N. -                                                                        GRUPO:                                            \n"
        header += " SETOR:                                                                               SUBGRUPO:                                         \n"
        header += " TIPO OS:                                                      STATUS: TODOS          MATERIAL:                                          \n"
        header += " ---------------------------------------------------------------------------------------------------------------------------------------\n"
        header += " Data       N.Ordem    Material                                                                       Quantidade Unidade           Valor\n"
        header += " ---------------------------------------------------------------------------------------------------------------------------------------\n"

        processed_lines.append(header)

        foi = 0
        templine = ''
        total = 0
        for line in lines:
            if "VEICULO: " in line and any(c.isdigit() for c in line.split("VEICULO: ")[1].split(" ")[0].strip()):
                if foi != 0:
                    # Adiciona a linha com o total do veículo anterior
                    if total > 0:
                        templine = templine.rstrip()
                        total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}\n'
                        processed_lines.append(total_line)

                # Reinicia o total para o novo veículo
                total = 0
                foi = 1
                templine = line
                processed_lines.append(templine)
                continue

            if not line.strip() or not line.strip()[0].isdigit() or "   OLEO" in line or "   ARLA" in line or "000" in line:
                continue

            processed_lines.append(line)

            # Extrai o valor da linha
            value_str = line.strip().split()[-1].replace('.', '').replace(',', '.')
            value = float(value_str)
            total += value

        # Adiciona o total do último veículo
        if total > 0:
            templine = templine.rstrip()
            total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}\n'
            processed_lines.append(total_line)

        processed_text = ''.join(processed_lines)

        with open('arquivo_processado.txt', 'w') as file:
            file.write(processed_text)

        # Remove as linhas em branco do arquivo processado
        remove_blank_lines('arquivo_processado.txt')

        # Cria um novo arquivo Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Estilo para o cabeçalho
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='000000')

        # Separa as colunas do cabeçalho
        header_columns = header.split('  ')

        # Adiciona o cabeçalho formatado no Excel
        for column in header_columns:
            header_cells = column.split('  ')
            row = []
            for cell_text in header_cells:
                cell = worksheet.cell(row=worksheet.max_row + 1, column=len(row) + 1, value=cell_text.strip())
                cell.font = header_font
                cell.fill = header_fill
                row.append(cell)

        # Abre o arquivo processado para leitura
        with open('arquivo_processado.txt', 'r') as file:
            processed_lines = file.readlines()

        # Adiciona as linhas processadas no Excel
        for line in processed_lines:
            row = line.split()
            worksheet.append(row)

        # Estilo para as células
        cell_font = Font(name='Arial', size=10)

        # Aplica o estilo às células
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = cell_font

        # Ajusta o tamanho da coluna "Total"
        worksheet.column_dimensions['E'].width = 15

        # Mescla as células da última linha (Total)
        worksheet.merge_cells(f"A{len(lines)+3}:D{len(lines)+3}")

        # Salva o arquivo Excel
        current_date = datetime.now().strftime("%m-%Y")
        output_file_name = f'arquivoExcel_{current_date}.xlsx'
        workbook.save(output_file_name)

    else:
        tk.messagebox.showinfo("Sem Arquivo", "NENHUM ARQUIVO SELECIONADO!")


def select_file():
    # Abre a janela de seleção de arquivo
    file_path = filedialog.askopenfilename()
    if file_path:
        # Chama a função para processar o arquivo selecionado
        process_file(file_path)

        # Mostra uma mensagem de conclusão
        tk.messagebox.showinfo("Concluído", "Arquivo processado e salvo como Excel com sucesso!")
    else:
        tk.messagebox.showinfo("Sem Arquivo", "NENHUM ARQUIVO SELECIONADO!")


# Cria uma janela principal
root = tk.Tk()
root.title("Limpador de Arquivos")

# Define o ícone da janela
icon = tk.PhotoImage(file="dependencias/icone.png")
root.iconphoto(True, icon)

# Define o tamanho mínimo da janela
root.minsize(400, 300)

# Cria o frame para centralizar o botão
frame = tk.Frame(root)
frame.pack(expand=True)

# Cria o botão
button = tk.Button(frame, text="Selecionar Arquivo", command=select_file)
button.pack(pady=10)

root.mainloop()
