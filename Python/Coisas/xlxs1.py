import tkinter as tk
from tkinter import filedialog
import codecs
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

def remove_blank_lines(file_path):
    # Read the content of the file
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # Remove blank lines
    non_blank_lines = [line for line in lines if line.strip()]

    # Write the processed content to a new file
    current_date = datetime.now().strftime("%m-%Y")
    output_file_name = f'arquivoLimpo_{current_date}.txt'
    with open(output_file_name, 'w') as file:
        file.writelines(non_blank_lines)

def process_file(file_path):
    if file_path:
        # Specify the correct file encoding
        encoding = 'UTF-8'  # Or another appropriate encoding

        # Open the file for reading using the correct encoding
        with codecs.open(file_path, 'r', encoding) as file:
            lines = file.readlines()

        processed_lines = []
        foi = 0
        templine = ''
        total = 0
        for line in lines:
            if "VEICULO: " in line and any(c.isdigit() for c in line.split("VEICULO: ")[1].split(" ")[0].strip()):
                if foi != 0:
                    # Add the line with the total of the previous vehicle
                    if total > 0:
                        templine = templine.rstrip()
                        total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}\n'
                        processed_lines.append(total_line)

                # Reset the total for the new vehicle
                total = 0
                foi = 1
                templine = line
                processed_lines.append(templine)
                continue

            if not line.strip() or not line.strip()[0].isdigit() or "   OLEO" in line or "   ARLA" in line or "000" in line:
                continue

            processed_lines.append(line)

            # Extract the value from the line
            value_str = line.strip().split()[-1].replace('.', '').replace(',', '.')
            value = float(value_str)
            total += value

        # Add the total of the last vehicle
        if total > 0:
            templine = templine.rstrip()
            total_line = ' TOTAL DO ' + templine + f': R$ {total:.2f}\n'
            processed_lines.append(total_line)

        processed_text = ''.join(processed_lines)

        with open('arquivo_processado.txt', 'w') as file:
            file.write(processed_text)

        # Remove blank lines from the processed file
        remove_blank_lines('arquivo_processado.txt')

        # Create a new Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Style for the header
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='000000')

        # Add the header formatted in Excel
        header_columns = ["Data", "N.Ordem", "Material", "Quantidade", "Unidade", "Valor"]
        header_row = []
        for column_text in header_columns:
            cell = worksheet.cell(row=1, column=len(header_row) + 1, value=column_text)
            cell.font = header_font
            cell.fill = header_fill
            header_row.append(cell)

        # Add the processed data to the Excel file
        with open('arquivo_processado.txt', 'r') as file:
            processed_lines = file.readlines()

        row_num = 2
        for line in processed_lines:
            columns = line.strip().split()
            for col_num, column_text in enumerate(columns, start=1):
                worksheet.cell(row=row_num, column=col_num, value=column_text)
            row_num += 1

        # Merge cells and add the total at the bottom
        last_column_letter = openpyxl.utils.get_column_letter(len(header_columns))
        total_range = f'A{row_num}:{last_column_letter}{row_num}'
        worksheet.merge_cells(total_range)
        total_cell = worksheet.cell(row=row_num, column=1, value="TOTAL")
        total_cell.font = Font(name='Arial', size=12, bold=True)
        total_cell.alignment = openpyxl.styles.Alignment(horizontal='right')

        # Save the Excel file
        current_date = datetime.now().strftime("%m-%Y")
        output_file_name = f'arquivo_processado_{current_date}.xlsx'
        workbook.save(output_file_name)

        # Show a completion message
        tk.messagebox.showinfo("Concluído", "Arquivo processado e salvo como Excel com sucesso!")
    else:
        tk.messagebox.showinfo("Sem Arquivo", "NENHUM ARQUIVO SELECIONADO!")


def select_file():
    # Open the file selection dialog
    file_path = filedialog.askopenfilename()
    if file_path:
        # Call the function to process the selected file
        process_file(file_path)

        # Show a completion message
        tk.messagebox.showinfo("Concluído", "Arquivo processado e salvo com sucesso como Excel!")
    else:
        tk.messagebox.showinfo("Sem Arquivo", "NENHUM ARQUIVO SELECIONADO!")


# Create the main window
root = tk.Tk()
root.title("Limpador de Arquivos")

# Define the window icon
icon = tk.PhotoImage(file="dependencias/icone.png")
root.iconphoto(True, icon)

# Set the minimum window size
root.minsize(400, 300)

# Create a frame to center the button
frame = tk.Frame(root)
frame.pack(expand=True)

# Create the button
button = tk.Button(frame, text="Selecionar Arquivo", command=select_file)
button.pack(pady=10)

root.mainloop()
